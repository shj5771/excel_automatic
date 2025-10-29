# 파일명: mark_duplicates_excel.py
# 사용법: python mark_duplicates_excel.py real_excel.xlsx
# (입력파일은 C:\Users\ST\Desktop\excel_program\input 안에 있어야 함)

import sys
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# ===== 0️⃣ 경로 설정 =====
base_dir = Path(r"C:\Users\ST\Desktop\excel_program")
input_dir = base_dir / "input"
output_dir = base_dir / "output/origin"
output_dir.mkdir(exist_ok=True)

# ===== 1️⃣ 명령행 인자 확인 =====
if len(sys.argv) < 2:
    print("Usage: python mark_duplicates_excel.py real_excel.xlsx")
    sys.exit(1)

file_name = sys.argv[1]
src = input_dir / file_name

if not src.exists():
    print(f"❌ 입력 파일을 찾을 수 없습니다: {src}")
    sys.exit(1)

# ===== 2️⃣ 기본 설정 =====
phone_col = '매칭용'  # 전화번호 들어있는 컬럼명

# ===== 3️⃣ 전화번호 정규화 함수 =====
def normalize_phone(x: str) -> str:
    if pd.isna(x):
        return ''
    return re.sub(r'\D', '', str(x))

print(f"📂 입력 파일: {src}")
print("📊 전화번호 중복 스캔 중...")

# ===== 4️⃣ 엑셀 로드 및 정규화 =====
df = pd.read_excel(src, dtype=str, engine="openpyxl").fillna('')

if phone_col not in df.columns:
    print(f"❌ '{phone_col}' 컬럼이 존재하지 않습니다. 실제 컬럼명 목록: {list(df.columns)}")
    sys.exit(1)

# 정규화전화번호 컬럼 추가
df['정규화전화번호'] = df[phone_col].map(normalize_phone)

# ===== 5️⃣ 중복 계산 (빈값 제외) =====
non_empty = df[df['정규화전화번호'].str.strip() != '']
phone_counts = non_empty['정규화전화번호'].value_counts().to_dict()

# 중복 횟수/여부 계산 (전화번호 없으면 0, False)
df['중복횟수'] = df['정규화전화번호'].map(lambda x: phone_counts.get(x, 0))
df['중복여부'] = df['중복횟수'] > 1

print(f"✅ 중복 계산 완료 (총 {len(non_empty):,}건 검사됨)")

# ===== 6️⃣ 원본 엑셀에 결과 반영 =====
print("🧩 원본 엑셀에 중복정보 추가 중...")

wb = load_workbook(src)
ws = wb.active

# 기존 마지막 열 뒤에 새 컬럼 추가
max_col = ws.max_column
ws.cell(row=1, column=max_col + 1, value="중복횟수")
ws.cell(row=1, column=max_col + 2, value="중복여부")

# 전화번호 매핑 딕셔너리 생성
phone_map = df.set_index(df.index)[['정규화전화번호', '중복횟수', '중복여부']].to_dict(orient='records')

count_filled = 0
for row in range(2, ws.max_row + 1):
    idx = row - 2  # df와 엑셀의 행 인덱스 맞추기
    if idx < len(df):
        ws.cell(row=row, column=max_col + 1, value=df.iloc[idx]['중복횟수'])
        ws.cell(row=row, column=max_col + 2, value=str(df.iloc[idx]['중복여부']))
        count_filled += 1

# ===== 7️⃣ 결과 저장 =====
out_path = output_dir / f"{src.stem}_with_duplicates.xlsx"
wb.save(out_path)
wb.close()

print(f"\n🎉 완료! 총 {count_filled:,}건의 중복정보 입력됨")
print(f"📁 저장 위치: {out_path}")
