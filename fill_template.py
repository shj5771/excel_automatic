# 파일명: fill_template_debug.py
# 사용법: python fill_template_debug.py

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import re

template_path = Path(r"C:\Users\ST\Desktop\excel_program\hangawon.xlsx")
csv_path      = Path(r"C:\Users\ST\Desktop\excel_program\123_only_unique.csv")
output_path   = Path(r"C:\Users\ST\Desktop\excel_program\output\hangawon_filled.xlsx")

# ===== 1) CSV 로드 & 헤더 정리 =====
print("📂 CSV 로드 중...")
df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig").fillna("")

# 헤더 양쪽 공백 제거 + 보이지 않는 문자 제거(제로폭, non-breaking space 등)
def clean_header(s: str) -> str:
    s = re.sub(r"[\u200b\u200c\u200d\uFEFF\u00A0]", "", s)  # 제로폭/불가시 문자 제거
    return s.strip()

df.columns = [clean_header(c) for c in df.columns]

print("🔎 CSV 컬럼 목록:", list(df.columns))

# ===== 2) 매핑할 실제 컬럼명(정확히 지정) =====
PHONE_COL = "전화번호"   # CSV의 "전화번호"
URL_COL   = "URL"       # CSV의 "URL"   (중요: #(URL1) 아님)

# 혹시 또 다를 수 있어 대비용 대체 패턴 (자동 탐색)
def find_col(name, patterns):
    if name in df.columns:
        return name
    for pat in patterns:
        for c in df.columns:
            if re.search(pat, c, flags=re.I):
                print(f"⚠️  '{name}'를 못찾아 '{c}'로 대체 매칭합니다.")
                return c
    raise KeyError(f"❌ '{name}' 컬럼을 찾지 못했습니다. 실제 헤더: {list(df.columns)}")

PHONE_COL = find_col(PHONE_COL, [r"전화\s*번호", r"휴대", r"tel|phone"])
URL_COL   = find_col(URL_COL,   [r"url", r"#\(url1\)", r"링크"])

# 디버그 샘플 출력
print(f"✅ 최종 매핑 → 전화번호: '{PHONE_COL}', URL: '{URL_COL}'")
print("   예시 1행 값  →", "전화번호:", df.iloc[0][PHONE_COL], "| URL:", df.iloc[0][URL_COL])

# ===== 3) 템플릿 열기 (시트 지정) =====
wb = load_workbook(template_path)
ws = wb["수신번호샘플"] if "수신번호샘플" in wb.sheetnames else wb.active
print("🗂  쓰는 시트:", ws.title)

# (선택) 기존 A/G 데이터 지우고 시작하고 싶으면 주석 해제
# for row in ws.iter_rows(min_row=2, min_col=1, max_col=7):
#     for cell in row:
#         cell.value = None

# ===== 4) 쓰기 =====
start_row = 2
for i, row in df.iterrows():
    phone = str(row[PHONE_COL]).strip()
    url   = str(row[URL_COL]).strip()
    ws[f"A{start_row + i}"] = phone     # A열 ← 전화번호
    ws[f"G{start_row + i}"] = url       # G열 ← URL
    if (i + 1) % 5000 == 0:
        print(f"  ↳ {i+1:,}건 채워넣는 중... (예: A{start_row+i}={phone}, G{start_row+i}={url})")

wb.save(output_path)
print(f"\n🎉 완료! 저장: {output_path}")
