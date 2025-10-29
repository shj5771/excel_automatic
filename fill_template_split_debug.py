# 파일명: fill_template_split_debug.py
# 사용법: python fill_template_split_debug.py

### 엑셀 템플릿에 CSV 데이터를 채우고, 행 수에 따라 여러 파일로 분할 저장하는 스크립트 ###
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import re
import math

template_path = Path(r"C:\Users\ST\Desktop\excel_program\input\hangawon.xlsx")
csv_path      = Path(r"C:\Users\ST\Desktop\excel_program\output\real_csv_only_unique.csv")
output_dir    = Path(r"C:\Users\ST\Desktop\excel_program\output\unique")
output_dir.mkdir(exist_ok=True)

CHUNK_SIZE = 1499  # ✅ 한 파일당 최대 1500행씩 저장

# ===== 1) CSV 로드 & 헤더 정리 =====
print("📂 CSV 로드 중...")
df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig").fillna("")

def clean_header(s: str) -> str:
    s = re.sub(r"[\u200b\u200c\u200d\uFEFF\u00A0]", "", s)
    return s.strip()

df.columns = [clean_header(c) for c in df.columns]
print("🔎 CSV 컬럼 목록:", list(df.columns))

# ===== 2) 매핑할 실제 컬럼명 =====
PHONE_COL = "매칭용"
URL_COL   = "조사 URL"

def find_col(name, patterns):
    if name in df.columns:
        return name
    for pat in patterns:
        for c in df.columns:
            if re.search(pat, c, flags=re.I):
                print(f"⚠️  '{name}'를 못찾아 '{c}'로 대체 매칭합니다.")
                return c
    raise KeyError(f"❌ '{name}' 컬럼을 찾지 못했습니다. 실제 헤더: {list(df.columns)}")

PHONE_COL = find_col(PHONE_COL, [r"전화\s*번호", r"휴대", r"tel|phone"])
URL_COL   = find_col(URL_COL,   [r"url", r"#\(url1\)", r"링크"])

print(f"✅ 최종 매핑 → 전화번호: '{PHONE_COL}', URL: '{URL_COL}'")
print("   예시 1행 값 →", "전화번호:", df.iloc[0][PHONE_COL], "| URL:", df.iloc[0][URL_COL])

# ===== 3) 분할 저장 =====
num_chunks = math.ceil(len(df) / CHUNK_SIZE)
print(f"\n📑 총 {len(df):,}행을 {num_chunks}개의 파일로 나눕니다 (파일당 {CHUNK_SIZE}행).")

for i in range(num_chunks):
    start_idx = i * CHUNK_SIZE
    end_idx = min((i + 1) * CHUNK_SIZE, len(df))
    sub_df = df.iloc[start_idx:end_idx]

    wb = load_workbook(template_path)
    ws = wb["수신번호샘플"] if "수신번호샘플" in wb.sheetnames else wb.active

    start_row = 2
    for j, row in sub_df.iterrows():
        phone = str(row[PHONE_COL]).strip()
        url = str(row[URL_COL]).strip()
        ws[f"A{start_row + (j - start_idx)}"] = phone
        ws[f"G{start_row + (j - start_idx)}"] = url

    # ✅ 마지막 파일이면 비어 있는 행 삭제
    if i == num_chunks - 1:
        print("🧹 마지막 템플릿에서 빈행 삭제 중...")
        max_row = ws.max_row
        for row in range(max_row, start_row, -1):  # 뒤에서부터 순회
            phone = ws[f"A{row}"].value
            url = ws[f"G{row}"].value
            if not phone and not url:
                ws.delete_rows(row)

    out_path = output_dir / f"hangawon_filled_part_{i+1}.xlsx"
    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path.name} ({start_idx+1:,} ~ {end_idx:,}행)")


print("\n🎉 전체 완료! 모든 파일이 output 폴더에 저장되었습니다.")
